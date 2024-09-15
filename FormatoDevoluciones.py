import pandas as pd
import os
from tkinter import Tk, filedialog
from openpyxl import load_workbook

"""
FormatoDevoluciones.py
Este script se utiliza para generar archivos CSV a partir de un archivo Excel que contiene información de artículos. 
El formato generado es utilizado en el proceso de actualización de precios de artículos en mi entorno de trabajo.

Uso:
1. El usuario selecciona un archivo Excel y una carpeta de destino.
2. Se lee cada hoja del Excel.
3. Se genera un CSV para cada hoja del Excel, con un formato predefinido.
4. El archivo CSV se guarda en la carpeta seleccionada.

El formato del DataFrame utilizado para generar los archivos CSV está adaptado al formato específico que se utiliza en mi 
trabajo para la actualización de precios. Este formato puede ser modificado según las necesidades particulares de cada usuario. 
Puedes ajustar las columnas y los datos en el DataFrame según los requisitos específicos de tu sistema o proceso.

"""

# Función para seleccionar un archivo Excel
def SeleccionarExcel():
    # Se abre un cuadro de diálogo para que el usuario seleccione el archivo Excel
    RutaExcel = filedialog.askopenfilename(title="Seleccionar Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
    return RutaExcel

# Función para seleccionar la carpeta de destino
def SeleccionarCarpeta():
    # Se abre un cuadro de diálogo para que el usuario seleccione la carpeta de destino para guardar los CSV
    RutaCarpeta = filedialog.askdirectory(title="Seleccionar carpeta de destino")
    return RutaCarpeta

# Función principal para procesar el Excel y generar los CSV
def ProcesarExcel(RutaExcel, RutaCarpeta):
    # Cargar el libro de Excel con openpyxl
    LibroTrabajo = load_workbook(RutaExcel)
    # Obtener los nombres de todas las hojas en el Excel
    NombresHojas = LibroTrabajo.sheetnames

    # Iterar sobre cada hoja en el archivo Excel
    for NombreHoja in NombresHojas:
        Hoja = LibroTrabajo[NombreHoja]

        # Crear un diccionario para almacenar los datos de la hoja actual en formato de columnas.
        # A continuación se especifica qué valores se colocarán en cada columna del archivo CSV que se va a generar.
        # Estos valores pueden ser ajustados según las necesidades del usuario.
        Datos = {
            'A': ['ParentKey', 'ItemCode'] + [Hoja[f'A{i}'].value for i in range(2, Hoja.max_row + 1)],  # Columna A: ParentKey, ItemCode, y datos de la columna A del Excel
            'B': ['LineNum', 'LineNum'] + [36] * (Hoja.max_row - 1),  # Columna B: Valor fijo 36 para todas las filas
            'C': ['PriceList', 'PriceList'] + [37] * (Hoja.max_row - 1),  # Columna C: Valor fijo 37 para todas las filas
            'D': ['Price', 'Price'] + [Hoja[f'F{i}'].value for i in range(2, Hoja.max_row + 1)],  # Columna D: Precios tomados de la columna F del Excel
            'E': ['Currency', 'Currency'] + ['$'] * (Hoja.max_row - 1),  # Columna E: Valor fijo "$" para todas las filas
        }

        # Crear un DataFrame con pandas usando el diccionario de datos
        DataFrame = pd.DataFrame(Datos)

        # Definir la ruta y el nombre del archivo CSV que se generará
        RutaCSV = os.path.join(RutaCarpeta, f"Devoluciones Palacio {NombreHoja}.csv")
        # Guardar el DataFrame como un archivo CSV, sin índices ni encabezados
        DataFrame.to_csv(RutaCSV, index=False, header=False)

    # Mensaje de confirmación cuando el proceso ha terminado
    print("Proceso completado.")

# Configurar la interfaz gráfica
def CrearInterfaz():
    # Crear una ventana oculta usando Tkinter
    VentanaPrincipal = Tk()
    VentanaPrincipal.withdraw()  # Ocultar la ventana principal ya que solo usaremos los cuadros de diálogo

    # Llamar a la función para seleccionar el archivo Excel
    RutaExcel = SeleccionarExcel()
    if not RutaExcel:
        # Si no se selecciona un archivo, mostrar un mensaje y salir
        print("No se seleccionó ningún archivo Excel.")
        return

    # Llamar a la función para seleccionar la carpeta de destino
    RutaCarpeta = SeleccionarCarpeta()
    if not RutaCarpeta:
        # Si no se selecciona una carpeta, mostrar un mensaje y salir
        print("No se seleccionó ninguna carpeta de destino.")
        return

    # Llamar a la función para procesar el archivo Excel y generar los CSV
    ProcesarExcel(RutaExcel, RutaCarpeta)

# Ejecutar la interfaz cuando se ejecuta el script directamente
if __name__ == "__main__":
    CrearInterfaz()
