import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

"""
FormatoActualizacionPrecios.py
Este script genera un archivo Excel que sigue un formato específico para la actualización de precios de artículos según la lista seleccionada.
Está diseñado para ser utilizado en mi entorno de trabajo para generar el formato que se utiliza para actualizar precios en
varias listas como 'Wholesale', 'Devoluciones', 'Amazon', etc.

Uso:
1. El usuario ingresa el mes, fecha y lista, los cuales se utilizan para acceder a la ruta donde se encuentra el archivo Excel.
2. El archivo Excel contiene la información necesaria para generar el formato de actualización de precios.
3. Se añade una segunda hoja en el mismo Excel, donde se generan los datos formateados para la actualización de precios.

El formato de las celdas y los datos generados puede ser adaptado según las necesidades de cada usuario. Puedes ajustar las fórmulas,
el contenido de las celdas y el formato de las hojas de acuerdo con los requisitos específicos de tu proceso o sistema.

"""

# Pedir al usuario ingresar datos para la ruta del archivo
Mes = input("Ingrese el mes: ")
Fecha = input("Ingrese la fecha: ")
Lista = input("Ingrese la lista: ")

# Cargar el archivo Excel desde la ruta generada con los datos ingresados
wb = openpyxl.load_workbook(f"Ejemplos/{Mes}/{Fecha}/{Lista} {Fecha}.xlsx")

Hoja1 = wb['Hoja1']

# Definir estilos de fuente, alineación y bordes
Negritas = Font(bold=True, size=10, name='Aptos Narrow')
Fuente = Font(size=10, name='Aptos Narrow')
Centrar = Alignment(horizontal="center")
Bordes = Border(top=Side(style='thin'),
                right=Side(style='thin'),
                bottom=Side(style='thin'),
                left=Side(style='thin'))

# Agregar encabezados en la hoja principal
Hoja1['D1'] = 'Concatenar'
Hoja1['C1'] = 'Buscar'

# Iterar sobre las filas para generar fórmulas y concatenar valores
for Fila in range(2, Hoja1.max_row + 1):
    ValorCelda = Hoja1[f'A{Fila}'].value
    Texto = f"'{ValorCelda}',"
    Hoja1[f'D{Fila}'] = Texto

    Formula = f"=VLOOKUP(A{Fila}, F:G, 2, FALSE)"
    Hoja1[f'C{Fila}'] = Formula

# Aplicar estilos de fuente, alineación y bordes a las celdas
for row in Hoja1.iter_rows():
    for cell in row:
        cell.font = Fuente
        cell.alignment = Centrar
        cell.border = Bordes

# Crear una segunda hoja para el formato de actualización de precios
wb.create_sheet('Hoja2')
Hoja2 = wb['Hoja2']

# Encabezados de la segunda hoja
Hoja2['A1'] = 'ParentKey'
Hoja2['A2'] = 'ItemCode'
Hoja2['B1'] = 'LineNum'
Hoja2['B2'] = 'LineNum'
Hoja2['C1'] = 'PriceList'
Hoja2['C2'] = 'PriceList'
Hoja2['D1'] = 'Price'
Hoja2['D2'] = 'Price'
Hoja2['E1'] = 'Currency'
Hoja2['E2'] = 'Currency'

# Copiar datos de la primera hoja a la segunda
UltimaFila = Hoja1.max_row
for fila in range(2, UltimaFila + 1):
    Hoja2[f'A{fila+1}'] = f'=Hoja1!C{fila}'
    Hoja2[f'D{fila+1}'] = f'=Hoja1!B{fila}'
    Hoja2[f'E{fila+1}'] = '$'

# Aplicar diferentes valores según la lista seleccionada
for fila in range(2, UltimaFila + 1):
    if Lista == 'Wholesale':
        Hoja2[f'B{fila+1}'] = 1
        Hoja2[f'C{fila+1}'] = 2
    elif Lista == 'Devoluciones':
        Hoja2[f'B{fila+1}'] = 30
        Hoja2[f'C{fila+1}'] = 31
    elif Lista == 'Salderos':
        Hoja2[f'B{fila+1}'] = 34
        Hoja2[f'C{fila+1}'] = 35
    elif Lista == 'Devoluciones Palacio':
        Hoja2[f'B{fila+1}'] = 36
        Hoja2[f'C{fila+1}'] = 37
    elif Lista == 'Amazon':
        Hoja2[f'B{fila+1}'] = 42
        Hoja2[f'C{fila+1}'] = 43
    elif Lista == 'Innova':
        Hoja2[f'B{fila+1}'] = 43
        Hoja2[f'C{fila+1}'] = 44
    elif Lista == 'Mercado Libre':
        Hoja2[f'B{fila+1}'] = 44
        Hoja2[f'C{fila+1}'] = 45

# Guardar el archivo Excel modificado
wb.save(f"Ejemplos/{Mes}/{Fecha}/{Lista} {Fecha}.xlsx")

# Confirmar que el proceso ha finalizado
print("¡Listo!")
